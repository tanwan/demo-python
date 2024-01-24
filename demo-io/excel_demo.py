import unittest
from openpyxl import load_workbook
import xlrd
import os
from pathlib import Path

curdir = Path(__file__).parent


class ElsxTest(unittest.TestCase):
    """07版使用openpyxl"""

    def setUp(self) -> None:
        # 支持相对路径
        self.wb = load_workbook(curdir / ".file/demo.xlsx")

    def test_sheet(self):
        """
        sheet的一些用法
        wb.sheetnames: 所有的sheet的名称
        wb.worksheets: 获取所有的sheet
        wb.get_sheet_by_name: 根据名称获取sheet(不存在报错)
        wb.index(sheet): 获取sheet的索引
        sheet.title: 名称
        sheet.max_row: 行数
        sheet.max_column: 列数
        sheet[row]获取行,1是第1行, row[col]获取列,0是第一列
        """
        # 所有的sheet的名称
        print("sheet names:", self.wb.sheetnames)

        # wb.worksheets获取所有的sheet,wb.get_sheet_by_name(name)根据名称(不存在报错),wb.index(sheet)获取sheet的索引
        sheet0 = self.wb.worksheets[0]

        # sheet[row]获取行,1是第1行, row[col]获取列,0是第一列
        print("cell00:", sheet0[1][0].value)

        # 遍历workbook可以得到sheet
        for sheet in self.wb:
            # sheet.title名称,sheet.max_row行数,sheet.max_column列数
            print("sheet name:", sheet.title, "row:", sheet.max_row, "col:", sheet.max_column)

    def test_row(self):
        """
        按row遍历
        for row in sheet｜sheet.rows
        for row in sheet.iter_rows(min_row=None, max_row=None, min_col=None, max_col=None) 可以指定遍历范围

        """
        sheet = self.wb.worksheets[0]
        # 等同于for row in sheet.rows
        for row in sheet:
            print(row)

        # 使用iter_rows可以限制遍历的范围
        for row in sheet.iter_rows(min_row=1, max_row=3, min_col=3, max_col=4):
            print(row)

    def test_column(self):
        """
        按列遍历
        for row in sheet.columns
        for row in sheet.iter_cols(min_row=None, max_row=None, min_col=None, max_col=None) 可以指定遍历范围
        """
        sheet = self.wb.worksheets[0]
        for col in sheet.rows:
            print(col)

        # 使用iter_rows可以限制遍历的范围
        for col in sheet.iter_cols(min_row=1, max_row=3, min_col=3, max_col=4):
            print(col)

    def test_cell(self):
        """
        cell的操作
        cell.col_idx: 索引
        cell.value: 值, 不存在则为None
        """
        sheet = self.wb.worksheets[0]
        for cell in sheet[1]:
            print(cell.col_idx, cell.value)


class XlsTest(unittest.TestCase):
    """03版使用xlrd"""

    def setUp(self) -> None:
        self.wb = xlrd.open_workbook(curdir / ".file/demo.xls")

    def test_sheet(self):
        """
        sheet的一些用法
        wb.sheet_names(): 所有sheet的名称
        wb.sheets(): 所有的sheet
        wb.sheet_by_index: 根据索引获取sheet(不存在报错)
        wb.sheet_by_name: 根据名称获取sheet(不存在报错)
        sheet.name: 名称
        sheet.nrows: 行数
        sheet.ncols: 列数
        sheet.row_values(n): 第n+1行所有的值的列表, 空值都是空字符串
        sheet.col_values(n): 第n+1列所有的值的列表, 空值都是空字符串
        sheet.row(n): 第n+1行
        sheet.col(n): 第n+1列
        sheet.cell(row, col): 第row+1行, 第col+1列的单元格
        """
        print("sheet names:", self.wb.sheet_names())

        # wb.sheets()获取所有的sheet,wb.sheet_by_index(0)根据索引(不存在报错),wb.sheet_by_name(name)根据名称(不存在报错)
        sheet0 = self.wb.sheet_by_index(0)

        # wb遍历得到sheet
        for sheet in self.wb:
            # sheet.name名称,sheet.nrows行数,sheet.ncols列数
            print("sheet name:", sheet.name, "rows:", sheet.nrows, "cols:", sheet.ncols)

        # sheet.row_values(n): 第n+1行所有的值的列表, sheet.col_values(n): 第n+1列所有的值的列表, 空值都是空字符串
        print("row0:", sheet0.row_values(0))
        print("col0:", sheet0.col_values(0))

        # sheet.row(n): 第n+1行
        print("row0:", sheet0.row(0))
        print("col0:", sheet0.col(0))
        # sheet.cell(row, col): 第row+1行, 第col+1列的单元格
        print("cell00:", sheet0.cell(0, 0))

    def test_row(self):
        """行遍历"""
        sheet = self.wb.sheet_by_index(0)
        for row in sheet:
            print(row)

        # 由于cell无法获取到索引, 所以有需要用的索引的话, 需要使用此方法遍历
        for r in range(0, sheet.nrows):
            row = sheet.row(r)
            for c in range(0, sheet.ncols):
                print("row:", r, "col:", c, "cell:", row[c])

    def test_cell(self):
        """
        cell的一些操作
        cell.value
        """
        for cell in self.wb.sheet_by_index(0).row(0):
            # 无法获取到cell的索引
            print(cell.value)
