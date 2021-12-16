import unittest
from openpyxl import load_workbook
import xlrd
import os
from pathlib import PurePath

curdir = PurePath(__file__).parent


class ExcelTest(unittest.TestCase):
    def test_read_xlsx(self):
        """07版本,使用openpyxl"""
        wb = load_workbook(f"{curdir}/file/demo.xlsx")
        # 所有的sheet的名称
        print(f"sheet names:{wb.sheetnames}")

        # wb.worksheets获取所有的sheet,wb.get_sheet_by_name(name)根据名称(不存在报错),wb.index(sheet)获取sheet的索引
        sheet0 = wb.worksheets[0]

        # sheet.title名称,sheet0.max_row行数,sheet.max_column列数
        print(f"sheet name:{sheet0.title},row:{sheet0.max_row},col:{sheet0.max_column}")

        # sheet[row]获取行,1是第1行, row[col]获取列,0是第一行
        print(f"cell00:{sheet0[1][0].value}")

        # 遍历
        for sheet in wb:
            # 使用rows, value不存在则为None
            for row in sheet.rows:
                for cell in row:
                    print(f"use rows, cell index:{cell.col_idx}, value:{cell.value}")

            # 使用iter_rows,可以通过min_row=None, max_row=None, min_col=None, max_col=None来限定遍历范围
            for row in sheet.iter_rows():
                for cell in row:
                    print(f"use iter_rows, cell index:{cell.col_idx}, value:{cell.value}")

            # 默认遍历
            for row in sheet:
                for cell in row:
                    # cell.col_idx列索引,cell.value值
                    print(f"cell index:{cell.col_idx}, value:{cell.value}")

            # 使用columns,按列遍历
            for columns in sheet.columns:
                for cell in columns:
                    print(f"use columns, cell index:{cell.col_idx}, value:{cell.value}")

            # 使用iter_cols,按列遍历,可以通过min_col=None, max_col=None, min_row=None, max_row=None来限定遍历范围
            for columns in sheet.iter_cols():
                for cell in columns:
                    print(f"use iter_cols, cell index:{cell.col_idx}, value:{cell.value}")

            # 使用values,只能获取到值
            for values in sheet.values:
                # values是一个tuple
                for value in values:
                    print(f"use values, cell value:{value}")

    def test_read_xls(self):
        """03版本,使用xlrd"""
        wb = xlrd.open_workbook(f"{curdir}/file/demo.xls")
        # wb.sheet_names()所有的sheet的名称
        print(f"sheet names:{wb.sheet_names()}")

        # wb.sheets()获取所有的sheet,wb.sheet_by_index(0)根据索引(不存在报错),wb.sheet_by_name(name)根据名称(不存在报错)
        sheet0 = wb.sheet_by_index(0)

        # sheet.name名称,sheet.nrows行数,sheet.ncols列数
        print(f"sheet name:{sheet0.name},rows:{sheet0.nrows},cols:{sheet0.ncols}")

        # sheet.row_values(0)第1行的所有值,sheet.col_values(0)第一列的所有值, 空值都是空字符串
        print(f"row0:{sheet0.row_values(0)},col0:{sheet0.col_values(0)}")

        # sheet.cell(0,0)用sheet获取cell, sheet.row(0)[0]sheet先获取row,再获取cell,sheet.col(0)[0]sheet先获取col,再获取cell
        print(f"cell00 by sheet:{sheet0.cell(0, 0).value}, cell00 by row:{sheet0.row(0)[0].value}, cell00 by col:{sheet0.col(0)[0].value}")

        # 遍历
        for sheet in wb:
            # 无索引的遍历
            for row in sheet:
                for cell in row:
                    # cell.value值, 没有提供索引的属性,所以获取不到cell的索引
                    print(f"cell  value:{cell.value}")

            # 有索引的遍历
            for r in range(0, sheet.nrows):
                row = sheet.row(r)
                # 使用cell
                for c in range(0, sheet.ncols):
                    print(f"use cell, row:{r}, cell index:{c},value:{row[c].value}")

                # 直接获取row_values
                for i, v in enumerate(sheet.row_values(r)):
                    print(f"use row_values, row:{r}, cell index:{i},value:{v}")
